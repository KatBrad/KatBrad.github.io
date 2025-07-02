import os
import pandas as pd
import numpy as np
from openpyxl import Workbook

# === Folder path with CSV files ===
folder_path = r"C:\Users\60107393\OneDrive - University of Doha for Science and Technology\Research\UREP\30th Cycle\IAT\April\Rename\Consent\Extract\step4"

# === Define blocks and output rows ===
blocks = {
    3: ('MnA1', 'SDA1', 124),
    4: ('MnA2', 'SDA2', 125),
    6: ('MnB1', 'SDB1', 126),
    7: ('MnB2', 'SDB2', 127),
}

# === Loop over each CSV file in the folder ===
for filename in os.listdir(folder_path):
    if filename.endswith('.csv') and not filename.startswith('~$'):
        csv_path = os.path.join(folder_path, filename)
        excel_path = os.path.join(folder_path, os.path.splitext(filename)[0] + '.xlsx')

        print(f"🔄 Processing {filename}...")

        # Load CSV
        df = pd.read_csv(csv_path)

        # Ensure E, J, and K columns are numeric
        df.iloc[:, 4] = pd.to_numeric(df.iloc[:, 4], errors='coerce')  # column E
        df.iloc[:, 9] = pd.to_numeric(df.iloc[:, 9], errors='coerce')  # column J
        df.iloc[:, 10] = pd.to_numeric(df.iloc[:, 10], errors='coerce')  # column K

        # Filter rows where column J == 0
        df_filtered = df[df.iloc[:, 9] == 0]

        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "IAT Stats"

        # Calculate mean and SD for each block
        for block_val, (mean_label, sd_label, row_num) in blocks.items():
            block_data = df_filtered[df_filtered.iloc[:, 10] == block_val].iloc[:, 4].dropna()

            if not block_data.empty:
                mean_val = block_data.mean()
                std_val = block_data.std()
            else:
                mean_val = std_val = np.nan

            ws[f"A{row_num}"] = mean_label
            ws[f"B{row_num}"] = mean_val
            ws[f"C{row_num}"] = sd_label
            ws[f"D{row_num}"] = std_val

        # Save Excel file
        wb.save(excel_path)
        print(f"✅ Saved Excel summary to {excel_path}")
